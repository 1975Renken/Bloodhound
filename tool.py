#!/usr/bin/env python3
"""
DeKalb County Meeting Minutes PDF Mining Tool
Searches government meeting PDFs for keywords related to governance and employee conduct issues
"""

import os
import re
import csv
import json
import time
import logging
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from urllib.parse import urljoin, urlparse
import requests
from bs4 import BeautifulSoup
import PyPDF2
import pdfplumber
import pytesseract
from PIL import Image
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from collections import defaultdict
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm
import hashlib

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pdf_miner.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class DeKalbPDFMiner:
    """Main class for mining DeKalb County meeting minutes PDFs"""
    
    def __init__(self, base_dir: str = "dekalb_pdfs"):
        """Initialize the PDF miner with configuration"""
        self.base_dir = base_dir
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        # Committee information
        self.committees = {
            'finance_administration': {
                'name': 'Finance & Administration',
                'main_url': 'https://dekalbcounty.org/government/county-boards-commissions/finance/',
                'archive_url': 'https://dekalbcounty.org/government/county-boards-commissions/finance/finance-committee-archive/'
            },
            'highway': {
                'name': 'Highway',
                'main_url': 'https://dekalbcounty.org/government/county-boards-commissions/county-highway/',
                'archive_url': 'https://dekalbcounty.org/government/county-boards-commissions/county-highway/highway-committee-archive/'
            },
            'law_justice': {
                'name': 'Law & Justice',
                'main_url': 'https://dekalbcounty.org/government/county-boards-commissions/law-justice/',
                'archive_url': 'https://dekalbcounty.org/government/county-boards-commissions/law-justice/law-justice-archive/'
            },
            'committee_whole': {
                'name': 'Committee of the Whole',
                'main_url': 'https://dekalbcounty.org/government/county-boards-commissions/committee-of-the-whole/',
                'archive_url': 'https://dekalbcounty.org/government/county-boards-commissions/committee-of-the-whole/committee-of-the-whole-archive/'
            },
            'executive': {
                'name': 'Executive',
                'main_url': 'https://dekalbcounty.org/government/county-boards-commissions/executive/',
                'archive_url': 'https://dekalbcounty.org/government/county-boards-commissions/executive/executive-committee-archive/'
            },
            'county_board': {
                'name': 'County Board',
                'main_url': 'https://dekalbcounty.org/government/county-boards-commissions/county-board-meetings/',
                'archive_url': 'https://dekalbcounty.org/government/county-boards-commissions/county-board-meetings/county-board-archives/'
            },
            'board_review': {
                'name': 'Board of Review',
                'main_url': 'https://dekalbcounty.org/departments/assessment-office/board-of-review/board-of-review-meetings/',
                'archive_url': 'https://dekalbcounty.org/departments/assessment-office/board-of-review/board-of-review-meetings/board-of-review-archives/'
            }
        }
        
        # Keywords by priority
        self.keywords = {
            'priority_1': {
                'terms': [
                    r'\bSteve\s+Hamm\b', r'\bS\.\s*Hamm\b', r'\bHamm\b',
                    r'\bethics\s+training\b',
                    r'\babuse\s+of\s+(authority|position)\b',
                    r'\bconflict\s+of\s+interest\b',
                    r'\bemployee\s+misconduct\b',
                    r'\bhighway\s+department\b.*?\b(complaint|incident|investigation)\b',
                    r'\bG-K\s+Broncos\b', r'\bBroncos\b',
                    r'\bKingston\s+Park\b',
                    r'\btrailer\s+removal\b'
                ],
                'color': 'FF0000'  # Red
            },
            'priority_2': {
                'terms': [
                    r'\bethics\b.*?\b(training|policy|violation)\b',
                    r'\bcode\s+of\s+conduct\b',
                    r'\bemployee\s+handbook\b',
                    r'\bdisciplinary\s+action\b',
                    r'\b(grievance|complaint)\b',
                    r'\binappropriate\s+use\b',
                    r'\bpersonal\s+use\b.*?\b(vehicle|position|authority)\b',
                    r'\bsheriff\b.*?\bhighway\b',
                    r'\b(intimidation|threatening)\b',
                    r'\bretaliation\b'
                ],
                'color': 'FFA500'  # Orange
            },
            'priority_3': {
                'terms': [
                    r'\boversight\b',
                    r'\baccountability\b',
                    r'\binternal\s+investigation\b',
                    r'\boutside\s+counsel\b',
                    r'\blitigation\s+hold\b',
                    r'\b(lawsuit|legal\s+action)\b',
                    r'\bsettlement\b',
                    r'\binsurance\s+claim\b',
                    r'\b(FOIA|freedom\s+of\s+information)\b',
                    r'\bpublic\s+comment\b.*?\b(complaint|concern)\b'
                ],
                'color': 'FFFF00'  # Yellow
            },
            'priority_4': {
                'terms': [
                    r'\btraining\s+budget\b',
                    r'\bprofessional\s+development\b',
                    r'\bmandatory\s+training\b',
                    r'\bcompliance\s+training\b',
                    r'\bharassment\s+training\b',
                    r'\bdiscrimination\b',
                    r'\bhostile\s+work\s+environment\b'
                ],
                'color': '00FF00'  # Green
            }
        }
        
        # Initialize results storage
        self.results = []
        self.pdf_cache = {}
        
    def setup_directories(self):
        """Create directory structure for storing PDFs"""
        for committee_key in self.committees:
            committee_dir = os.path.join(self.base_dir, committee_key)
            os.makedirs(committee_dir, exist_ok=True)
            os.makedirs(os.path.join(committee_dir, 'current'), exist_ok=True)
            os.makedirs(os.path.join(committee_dir, 'archive'), exist_ok=True)
        
        # Create results directory
        os.makedirs('results', exist_ok=True)
        logger.info(f"Directory structure created under {self.base_dir}")
    
    def extract_pdf_links(self, url: str) -> List[Dict]:
        """Extract all PDF links from a webpage"""
        pdf_links = []
        try:
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find all links to PDFs
            for link in soup.find_all('a', href=True):
                href = link['href']
                if href.lower().endswith('.pdf'):
                    full_url = urljoin(url, href)
                    link_text = link.get_text(strip=True)
                    
                    # Try to extract date from filename or link text
                    date_match = re.search(r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})|(\w+\s+\d{1,2},?\s+\d{4})', link_text + ' ' + href)
                    date_str = date_match.group() if date_match else 'Unknown'
                    
                    pdf_links.append({
                        'url': full_url,
                        'text': link_text,
                        'date_str': date_str,
                        'filename': os.path.basename(urlparse(full_url).path)
                    })
            
            logger.info(f"Found {len(pdf_links)} PDFs on {url}")
        except Exception as e:
            logger.error(f"Error extracting PDFs from {url}: {e}")
        
        return pdf_links
    
    def download_pdf(self, pdf_info: Dict, save_path: str) -> bool:
        """Download a single PDF file"""
        try:
            # Check if already downloaded
            if os.path.exists(save_path):
                logger.info(f"PDF already exists: {save_path}")
                return True
            
            response = self.session.get(pdf_info['url'], timeout=60, stream=True)
            response.raise_for_status()
            
            # Save PDF
            with open(save_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            logger.info(f"Downloaded: {pdf_info['filename']}")
            return True
        except Exception as e:
            logger.error(f"Error downloading {pdf_info['url']}: {e}")
            return False
    
    def extract_text_from_pdf(self, pdf_path: str) -> List[Tuple[int, str]]:
        """Extract text from PDF, returns list of (page_num, text) tuples"""
        text_pages = []
        
        try:
            # Try with pdfplumber first (better for tables and complex layouts)
            with pdfplumber.open(pdf_path) as pdf:
                for i, page in enumerate(pdf.pages, 1):
                    text = page.extract_text()
                    if text and len(text.strip()) > 50:
                        text_pages.append((i, text))
                        
            # If no text extracted, try PyPDF2
            if not text_pages:
                with open(pdf_path, 'rb') as file:
                    reader = PyPDF2.PdfReader(file)
                    for i, page in enumerate(reader.pages, 1):
                        text = page.extract_text()
                        if text and len(text.strip()) > 50:
                            text_pages.append((i, text))
            
            # If still no text, might be scanned - try OCR
            if not text_pages:
                logger.info(f"No text found in {pdf_path}, attempting OCR...")
                text_pages = self.ocr_pdf(pdf_path)
                
        except Exception as e:
            logger.error(f"Error extracting text from {pdf_path}: {e}")
        
        return text_pages
    
    def ocr_pdf(self, pdf_path: str) -> List[Tuple[int, str]]:
        """OCR a scanned PDF using pytesseract"""
        text_pages = []
        try:
            import pdf2image
            
            # Convert PDF pages to images
            images = pdf2image.convert_from_path(pdf_path)
            
            for i, image in enumerate(images, 1):
                # Perform OCR
                text = pytesseract.image_to_string(image)
                if text and len(text.strip()) > 50:
                    text_pages.append((i, text))
                    
            logger.info(f"OCR completed for {pdf_path}: {len(text_pages)} pages")
        except Exception as e:
            logger.error(f"OCR failed for {pdf_path}: {e}")
        
        return text_pages
    
    def search_keywords_in_text(self, text: str, page_num: int, pdf_info: Dict, committee: str) -> List[Dict]:
        """Search for keywords in text and return matches with context"""
        matches = []
        text_lower = text.lower()
        
        for priority, priority_data in self.keywords.items():
            for pattern in priority_data['terms']:
                # Search with regex (case-insensitive)
                for match in re.finditer(pattern, text, re.IGNORECASE | re.DOTALL):
                    # Get context (50 words before and after)
                    start = max(0, match.start() - 300)
                    end = min(len(text), match.end() + 300)
                    context = text[start:end].strip()
                    
                    # Clean up context
                    context = ' '.join(context.split())
                    
                    matches.append({
                        'committee': committee,
                        'pdf_filename': pdf_info['filename'],
                        'pdf_date': pdf_info['date_str'],
                        'page': page_num,
                        'priority': priority,
                        'keyword_pattern': pattern,
                        'matched_text': match.group(),
                        'context': context,
                        'url': pdf_info['url']
                    })
        
        return matches
    
    def process_committee(self, committee_key: str):
        """Process all PDFs for a single committee"""
        committee = self.committees[committee_key]
        committee_name = committee['name']
        logger.info(f"\nProcessing {committee_name}...")
        
        # Get PDFs from main page
        current_pdfs = self.extract_pdf_links(committee['main_url'])
        
        # Get PDFs from archive page
        archive_pdfs = self.extract_pdf_links(committee['archive_url'])
        
        # Download and process current PDFs
        for pdf_info in tqdm(current_pdfs, desc=f"{committee_name} - Current"):
            save_path = os.path.join(self.base_dir, committee_key, 'current', pdf_info['filename'])
            if self.download_pdf(pdf_info, save_path):
                # Extract and search text
                text_pages = self.extract_text_from_pdf(save_path)
                for page_num, text in text_pages:
                    matches = self.search_keywords_in_text(text, page_num, pdf_info, committee_name)
                    self.results.extend(matches)
        
        # Download and process archive PDFs
        for pdf_info in tqdm(archive_pdfs, desc=f"{committee_name} - Archive"):
            save_path = os.path.join(self.base_dir, committee_key, 'archive', pdf_info['filename'])
            if self.download_pdf(pdf_info, save_path):
                # Extract and search text
                text_pages = self.extract_text_from_pdf(save_path)
                for page_num, text in text_pages:
                    matches = self.search_keywords_in_text(text, page_num, pdf_info, committee_name)
                    self.results.extend(matches)
    
    def generate_excel_report(self):
        """Generate comprehensive Excel report with multiple sheets"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = f'results/dekalb_findings_{timestamp}.xlsx'
        
        # Create DataFrame from results
        df = pd.DataFrame(self.results)
        
        if df.empty:
            logger.warning("No results to report")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: All Findings
            df.to_excel(writer, sheet_name='All Findings', index=False)
            
            # Sheet 2: Summary by Committee
            committee_summary = df.groupby(['committee', 'priority']).size().unstack(fill_value=0)
            committee_summary.to_excel(writer, sheet_name='Committee Summary')
            
            # Sheet 3: Summary by Priority
            priority_summary = df.groupby(['priority', 'committee']).size().unstack(fill_value=0)
            priority_summary.to_excel(writer, sheet_name='Priority Summary')
            
            # Sheet 4: Timeline Analysis
            if 'pdf_date' in df.columns:
                timeline_df = df.groupby(['pdf_date', 'committee']).size().unstack(fill_value=0)
                timeline_df.to_excel(writer, sheet_name='Timeline')
            
            # Sheet 5: Keyword Frequency
            keyword_freq = df['keyword_pattern'].value_counts()
            keyword_freq.to_excel(writer, sheet_name='Keyword Frequency')
            
            # Format the Excel file
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Add headers formatting
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Color-code priority cells in All Findings sheet
                if sheet_name == 'All Findings' and 'priority' in df.columns:
                    priority_col = None
                    for idx, col in enumerate(worksheet[1], 1):
                        if col.value == 'priority':
                            priority_col = idx
                            break
                    
                    if priority_col:
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=priority_col)
                            priority = cell.value
                            if priority in self.keywords:
                                color = self.keywords[priority]['color']
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        logger.info(f"Excel report generated: {excel_path}")
        return excel_path
    
    def generate_visualizations(self):
        """Generate visualization charts"""
        if not self.results:
            logger.warning("No results to visualize")
            return
        
        df = pd.DataFrame(self.results)
        
        # Set up the plot style
        plt.style.use('seaborn-v0_8-darkgrid')
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        
        # 1. Findings by Committee
        committee_counts = df['committee'].value_counts()
        axes[0, 0].bar(committee_counts.index, committee_counts.values)
        axes[0, 0].set_title('Total Findings by Committee')
        axes[0, 0].set_xlabel('Committee')
        axes[0, 0].set_ylabel('Number of Findings')
        axes[0, 0].tick_params(axis='x', rotation=45)
        
        # 2. Findings by Priority
        priority_counts = df['priority'].value_counts()
        colors = [self.keywords[p]['color'] for p in priority_counts.index if p in self.keywords]
        axes[0, 1].pie(priority_counts.values, labels=priority_counts.index, autopct='%1.1f%%', colors=['#' + c for c in colors])
        axes[0, 1].set_title('Findings by Priority Level')
        
        # 3. Top Keywords Found
        keyword_counts = df['matched_text'].value_counts().head(10)
        axes[1, 0].barh(keyword_counts.index, keyword_counts.values)
        axes[1, 0].set_title('Top 10 Keywords Found')
        axes[1, 0].set_xlabel('Frequency')
        
        # 4. Heatmap of Committee vs Priority
        pivot = df.pivot_table(index='committee', columns='priority', aggfunc='size', fill_value=0)
        sns.heatmap(pivot, annot=True, fmt='d', cmap='YlOrRd', ax=axes[1, 1])
        axes[1, 1].set_title('Committee vs Priority Heatmap')
        
        plt.tight_layout()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        chart_path = f'results/dekalb_analysis_{timestamp}.png'
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"Visualizations saved: {chart_path}")
        return chart_path
    
    def generate_summary_report(self):
        """Generate a text summary report"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_path = f'results/dekalb_summary_{timestamp}.txt'
        
        with open(summary_path, 'w') as f:
            f.write("=" * 80 + "\n")
            f.write("DEKALB COUNTY MEETING MINUTES ANALYSIS SUMMARY\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")
            
            if not self.results:
                f.write("No findings identified in the analyzed documents.\n")
                return summary_path
            
            df = pd.DataFrame(self.results)
            
            # Overall statistics
            f.write("OVERALL STATISTICS\n")
            f.write("-" * 40 + "\n")
            f.write(f"Total findings: {len(df)}\n")
            f.write(f"Committees analyzed: {df['committee'].nunique()}\n")
            f.write(f"PDFs processed: {df['pdf_filename'].nunique()}\n")
            f.write(f"Unique keywords matched: {df['keyword_pattern'].nunique()}\n\n")
            
            # Priority breakdown
            f.write("FINDINGS BY PRIORITY\n")
            f.write("-" * 40 + "\n")
            for priority in ['priority_1', 'priority_2', 'priority_3', 'priority_4']:
                count = len(df[df['priority'] == priority])
                percentage = (count / len(df)) * 100
                f.write(f"{priority}: {count} ({percentage:.1f}%)\n")
            f.write("\n")
            
            # Committee breakdown
            f.write("FINDINGS BY COMMITTEE\n")
            f.write("-" * 40 + "\n")
            for committee, count in df['committee'].value_counts().items():
                f.write(f"{committee}: {count}\n")
            f.write("\n")
            
            # High priority findings details
            f.write("HIGH PRIORITY (PRIORITY 1) FINDINGS\n")
            f.write("-" * 40 + "\n")
            priority_1 = df[df['priority'] == 'priority_1']
            if not priority_1.empty:
                for idx, row in priority_1.iterrows():
                    f.write(f"\nCommittee: {row['committee']}\n")
                    f.write(f"Date: {row['pdf_date']}\n")
                    f.write(f"Page: {row['page']}\n")
                    f.write(f"Keyword: {row['matched_text']}\n")
                    f.write(f"Context: ...{row['context'][:200]}...\n")
                    f.write("-" * 20 + "\n")
            else:
                f.write("No Priority 1 findings identified.\n")
            
            # Patterns and observations
            f.write("\n\nPATTERNS AND OBSERVATIONS\n")
            f.write("-" * 40 + "\n")
            
            # Check for committees with no findings
            all_committees = set(self.committees.keys())
            committees_with_findings = set(df['committee'].unique())
            committees_without = all_committees - committees_with_findings
            if committees_without:
                f.write(f"Committees with no keyword matches: {', '.join(committees_without)}\n")
            
            # Most common keywords
            f.write("\nMost frequently matched keywords:\n")
            for keyword, count in df['matched_text'].value_counts().head(5).items():
                f.write(f"  - '{keyword}': {count} occurrences\n")
        
        logger.info(f"Summary report generated: {summary_path}")
        return summary_path
    
    def run(self):
        """Main execution method"""
        logger.info("Starting DeKalb County PDF Mining Tool")
        logger.info("=" * 60)
        
        # Setup
        self.setup_directories()
        
        # Process each committee
        for committee_key in self.committees:
            self.process_committee(committee_key)
            time.sleep(2)  # Be respectful to the server
        
        # Generate reports
        logger.info("\nGenerating reports...")
        excel_report = self.generate_excel_report()
        chart_report = self.generate_visualizations()
        summary_report = self.generate_summary_report()
        
        # Final summary
        logger.info("\n" + "=" * 60)
        logger.info("MINING COMPLETE")
        logger.info(f"Total findings: {len(self.results)}")
        logger.info(f"Reports generated:")
        logger.info(f"  - Excel: {excel_report}")
        logger.info(f"  - Charts: {chart_report}")
        logger.info(f"  - Summary: {summary_report}")
        logger.info("=" * 60)
        
        return {
            'total_findings': len(self.results),
            'excel_report': excel_report,
            'chart_report': chart_report,
            'summary_report': summary_report,
            'results': self.results
        }


def main():
    """Main entry point"""
    miner = DeKalbPDFMiner()
    results = miner.run()
    
    # Print quick summary to console
    print("\n" + "=" * 60)
    print("DEKALB COUNTY PDF MINING COMPLETE")
    print("=" * 60)
    print(f"Total findings: {results['total_findings']}")
    
    if results['total_findings'] > 0:
        df = pd.DataFrame(results['results'])
        print("\nTop findings by priority:")
        for priority in ['priority_1', 'priority_2', 'priority_3', 'priority_4']:
            count = len(df[df['priority'] == priority])
            if count > 0:
                print(f"  {priority}: {count} findings")
        
        print("\nReports saved to 'results' folder")
        print("Check the Excel file for detailed findings and context")
    else:
        print("No keyword matches found in the analyzed documents.")
    
    return results


if __name__ == "__main__":
    main()